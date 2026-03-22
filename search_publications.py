"""
Поиск публикаций в ИС «Метафора» по DOI или названию и извлечение аннотаций и списка литературы.
Используется в mtfr_updater для кнопки «Подтянуть данные из MTFR».
Отдельно: чтение списка из input_files/publications_to_find.xlsx и сохранение в output_files/publications_result.csv (в т.ч. колонка «Страницы» из #dataTable_pubdata).
"""

import argparse
import csv
import io
import logging
import re
import sys
from pathlib import Path

from bs4 import BeautifulSoup

from metafora_client import MetaforaClient

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

PUBLICATIONS_URL = "/publications"
INPUT_XLSX = Path("input_files") / "publications_to_find.xlsx"
DEFAULT_OUTPUT_CSV = Path("output_files") / "publications_result.csv"
CSV_DELIMITER = ";"


def read_publications_from_xlsx(path: Path) -> list[tuple[str, str]]:
    """
    Читает список публикаций для поиска: пары (doi, название).
    Требует openpyxl. DOI может быть пустым — тогда поиск только по названию.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("Для read_publications_from_xlsx нужен openpyxl: pip install openpyxl")
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    doi_col = None
    title_col = None
    for i, h in enumerate(header):
        if h in ("doi", "дои"):
            doi_col = i
        if h in ("название", "title", "заголовок"):
            title_col = i
    if doi_col is None:
        doi_col = 0
    if title_col is None and len(header) > 1:
        title_col = 1
    result = []
    for row in rows[1:]:
        if not row:
            continue
        doi = ""
        if doi_col is not None and doi_col < len(row) and row[doi_col] is not None:
            doi = str(row[doi_col]).strip()
        title = ""
        if title_col is not None and title_col < len(row) and row[title_col] is not None:
            title = str(row[title_col]).strip()
        if not doi and not title:
            continue
        result.append((doi, title))
    return result


def search_publications_page(
    client: MetaforaClient,
    filter_doi: str = "",
    filter_title: str = "",
) -> str:
    """Запрашивает страницу /publications с фильтром по DOI и/или названию."""
    url = client._url(PUBLICATIONS_URL)
    params = {}
    if filter_doi:
        params["filterDoi"] = filter_doi
    if filter_title:
        params["filterTitle"] = filter_title
    r = client.session.get(url, params=params, timeout=client.timeout)
    r.raise_for_status()
    return r.text


def parse_publication_links(html: str, base_url: str) -> list[dict]:
    """
    Извлекает из страницы списка публикаций ссылки на карточки публикаций.
    Ищет <a href=".../publications/{uuid}" class="text-primary font-weight-bold"> и текст (название).
    """
    soup = BeautifulSoup(html, "html.parser")
    result = []
    for a in soup.find_all("a", href=True, class_=lambda c: c and "text-primary" in c and "font-weight-bold" in c):
        href = a.get("href", "")
        if "/publications/" not in href or href.rstrip("/").endswith("/publications"):
            continue
        full_url = href if href.startswith("http") else f"{base_url.rstrip('/')}{href}"
        title = (a.get_text() or "").strip()
        result.append({"url": full_url, "title": title})
    if not result:
        for a in soup.find_all("a", href=True):
            href = a.get("href", "")
            m = re.search(r"/publications/([a-f0-9-]{36})", href, re.I)
            if m:
                full_url = href if href.startswith("http") else f"{base_url.rstrip('/')}{href}"
                title = (a.get_text() or "").strip()
                if full_url not in [x["url"] for x in result]:
                    result.append({"url": full_url, "title": title})
    return result


def _section_text(soup: BeautifulSoup, headings: list[str]) -> str:
    """Ищет на странице секцию по одному из заголовков и возвращает текст до следующей секции."""
    for h in headings:
        tag = soup.find(
            ["h1", "h2", "h3", "h4", "h5", "h6", "strong", "b", "dt"],
            string=re.compile(re.escape(h), re.I),
        )
        if not tag:
            continue
        parts = []
        for s in tag.find_next_siblings():
            if s.name and s.name in ("h1", "h2", "h3", "h4", "h5", "h6"):
                break
            text = s.get_text(separator=" ", strip=True)
            if text:
                parts.append(text)
        return " ".join(parts).strip()
    return ""


def _find_list_or_paragraphs(soup: BeautifulSoup, headings: list[str]) -> str:
    """Ищет секцию (например, список литературы) и возвращает текст списка или параграфов."""
    for h in headings:
        tag = soup.find(
            ["h1", "h2", "h3", "h4", "h5", "h6", "strong", "b"],
            string=re.compile(re.escape(h), re.I),
        )
        if not tag:
            continue
        block = tag.find_next_sibling()
        items = []
        if block and block.name == "ol":
            for li in block.find_all("li"):
                items.append(li.get_text(separator=" ", strip=True))
        elif block and block.name == "ul":
            for li in block.find_all("li"):
                items.append(li.get_text(separator=" ", strip=True))
        else:
            for s in tag.find_next_siblings():
                if s.name and s.name in ("h1", "h2", "h3", "h4", "h5", "h6"):
                    break
                items.append(s.get_text(separator=" ", strip=True))
        if items:
            return "\n".join(items).strip()
    return ""


def _text_from_multilang_block(block) -> str:
    """Текст из блока .multilang без подписи языка (.lang_caption)."""
    cap_el = block.find(class_="lang_caption")
    full_text = block.get_text(separator=" ", strip=True)
    if not cap_el:
        return full_text
    cap_clean = cap_el.get_text(strip=True)
    cap_upper = cap_clean.upper()
    if full_text.upper().startswith(cap_upper):
        return full_text[len(cap_clean):].strip()
    return full_text


def _parse_all_multilang(soup: BeautifulSoup) -> list[tuple[str, str]]:
    """
    Извлекает все блоки .multilang в порядке появления.
    Возвращает список пар (lang, text), где lang — "ru" или "en".
    """
    result = []
    for block in soup.find_all(class_="multilang"):
        cap_el = block.find(class_="lang_caption")
        if not cap_el:
            continue
        cap = cap_el.get_text(strip=True).upper()
        value = _text_from_multilang_block(block)
        if not value:
            continue
        if "ENG" in cap or cap == "EN":
            result.append(("en", value))
        elif "РУС" in cap or "RUS" in cap or cap == "RU":
            result.append(("ru", value))
    return result


def _pubdata_row_cell_text(tr) -> str:
    """Текст ячеек данных в строке таблицы сведений о публикации (без заголовка <th>)."""
    tds = tr.find_all("td")
    if not tds:
        return ""
    parts: list[str] = []
    for td in tds:
        parts.append(td.get_text(separator=" ", strip=True))
    return " ".join(p for p in parts if p).strip()


def _parse_optional_positive_int(raw: str) -> int | None:
    s = (raw or "").strip()
    if not s or s in ("—", "–", "-"):
        return None
    if re.fullmatch(r"\d+", s):
        return int(s)
    return None


def _format_pages_field(first_page: int | None, last_page: int | None) -> str:
    if first_page is not None and last_page is not None:
        if first_page == last_page:
            return str(first_page)
        return f"{first_page}-{last_page}"
    if first_page is not None:
        return str(first_page)
    if last_page is not None:
        return str(last_page)
    return ""


def _parse_pubdata_pages(soup: BeautifulSoup) -> tuple[int | None, int | None, str]:
    """
    Извлекает первую/последнюю страницу из таблицы #dataTable_pubdata
    (блок «Данные публикации» на карточке в ИС Метафора).
    """
    table = soup.find("table", id="dataTable_pubdata")
    if not table:
        return None, None, ""
    first_page: int | None = None
    last_page: int | None = None
    for tr in table.find_all("tr"):
        th = tr.find("th")
        if not th:
            continue
        label = (th.get_text() or "").strip()
        cell = _pubdata_row_cell_text(tr)
        if label == "Первая страница":
            first_page = _parse_optional_positive_int(cell)
        elif label == "Последняя страница":
            last_page = _parse_optional_positive_int(cell)
    pages = _format_pages_field(first_page, last_page)
    return first_page, last_page, pages


def _parse_references_from_table(soup: BeautifulSoup) -> tuple[str, str]:
    """
    Извлекает списки литературы (RU/EN) из таблицы #dataTable_refs.
    """
    table = soup.find("table", id="dataTable_refs")
    if not table:
        for t in soup.find_all("table"):
            if t.find("th", string=re.compile(r"Список литературы|Литература|References", re.I)):
                table = t
                break
    if not table:
        return "", ""
    references_ru = ""
    references_en = ""
    for tr in table.find_all("tr"):
        th = tr.find("th")
        if th and not re.search(r"Список литературы|Литература|References", th.get_text() or "", re.I):
            continue
        for td in tr.find_all("td", class_="multilang"):
            cap_el = td.find(class_="lang_caption")
            if not cap_el:
                continue
            cap = cap_el.get_text(strip=True).upper()
            full = td.get_text(separator="\n", strip=True)
            cap_clean = cap_el.get_text(strip=True)
            value = full[len(cap_clean):].strip() if full.upper().startswith(cap_clean.upper()) else full
            if "ENG" in cap or cap == "EN":
                references_en = value
            elif "РУС" in cap or "RUS" in cap or cap == "RU":
                references_ru = value
        if references_ru or references_en:
            break
    return references_ru, references_en


def parse_publication_detail(html: str) -> dict:
    """
    Парсит страницу публикации: название (RU/EN), аннотация (RU), аннотация (EN), список литературы,
    номера страниц из таблицы #dataTable_pubdata (поле pages для формы — диапазон «478-492»).
    """
    soup = BeautifulSoup(html, "html.parser")
    first_page, last_page, pages = _parse_pubdata_pages(soup)
    multilang = _parse_all_multilang(soup)
    ru_texts = [t for lang, t in multilang if lang == "ru"]
    en_texts = [t for lang, t in multilang if lang == "en"]
    title_ru = ru_texts[0] if len(ru_texts) > 0 else ""
    title_en = en_texts[0] if len(en_texts) > 0 else ""
    abstract_ru = ru_texts[1] if len(ru_texts) > 1 else _section_text(soup, ["Аннотация", "Аннотация (рус)", "Резюме"])
    abstract_en = en_texts[1] if len(en_texts) > 1 else _section_text(soup, ["Abstract", "Аннотация (англ)", "Summary", "Аннотация на английском"])
    references_ru, references_en = _parse_references_from_table(soup)
    if not references_ru:
        references_ru = _find_list_or_paragraphs(
            soup,
            ["Список литературы", "Литература", "References", "Библиографический список"],
        )
        if not references_ru:
            references_ru = _section_text(soup, ["Список литературы", "Литература", "References"])
    return {
        "title_ru": title_ru,
        "title_en": title_en,
        "abstract_ru": abstract_ru,
        "abstract_en": abstract_en,
        "references_ru": references_ru,
        "references_en": references_en,
        "first_page": first_page,
        "last_page": last_page,
        "pages": pages,
    }


def fetch_publication_detail(client: MetaforaClient, url: str) -> str:
    """Загружает HTML страницы публикации по URL."""
    r = client.session.get(url, timeout=client.timeout)
    r.raise_for_status()
    return r.text


def main() -> None:
    """CLI: поиск по XLSX и сохранение в CSV. Требует openpyxl."""
    parser = argparse.ArgumentParser(
        description="Поиск публикаций в ИС Метафора по DOI/названию и извлечение аннотаций и списка литературы",
    )
    parser.add_argument(
        "-o", "--output",
        type=Path,
        default=None,
        help=f"Путь к CSV результата (по умолчанию: {DEFAULT_OUTPUT_CSV})",
    )
    parser.add_argument(
        "-q", "--quiet",
        action="store_true",
        help="Не выводить таблицу в консоль",
    )
    args = parser.parse_args()
    output_path = args.output or (Path(__file__).resolve().parent / DEFAULT_OUTPUT_CSV)
    output_path = Path(output_path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    input_path = Path(__file__).resolve().parent / INPUT_XLSX
    try:
        publications = read_publications_from_xlsx(input_path)
    except FileNotFoundError as e:
        logger.error("%s", e)
        sys.exit(1)
    if not publications:
        logger.warning("В файле %s не найдено ни одной записи (DOI или Название).", input_path)
        return

    client = MetaforaClient()
    if not client.login():
        logger.error("Не удалось войти в ИС Метафора.")
        sys.exit(1)

    headers = (
        "DOI (запрос)",
        "Название (запрос)",
        "URL публикации",
        "Название (RU)",
        "Название (EN)",
        "Аннотация (RU)",
        "Аннотация (EN)",
        "Страницы",
        "Список литературы (RU)",
        "Список литературы (EN)",
    )

    def write_row(row: list) -> None:
        writer.writerow(row)
        if not args.quiet:
            buf = io.StringIO()
            csv.writer(buf, delimiter=CSV_DELIMITER, quoting=csv.QUOTE_MINIMAL).writerow(row)
            print(buf.getvalue().rstrip("\r\n"))

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f_out:
        writer = csv.writer(f_out, delimiter=CSV_DELIMITER, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        if not args.quiet:
            buf = io.StringIO()
            csv.writer(buf, delimiter=CSV_DELIMITER, quoting=csv.QUOTE_MINIMAL).writerow(headers)
            print(buf.getvalue().rstrip("\r\n"))

        for doi, title in publications:
            try:
                html = search_publications_page(client, filter_doi=doi, filter_title=title)
                links = parse_publication_links(html, client.base_url)
            except Exception as e:
                logger.exception("Ошибка при поиске DOI=%s title=%s: %s", doi, title, e)
                write_row([doi, title, "", "", "", "", "", "", "", str(e)])
                continue

            if not links:
                write_row([doi, title, "", "не найдено", "", "", "", "", "", ""])
                continue

            pub = links[0]
            pub_url = pub["url"]
            try:
                detail_html = fetch_publication_detail(client, pub_url)
                detail = parse_publication_detail(detail_html)
            except Exception as e:
                logger.exception("Ошибка при загрузке страницы публикации %s: %s", pub_url, e)
                write_row([doi, title, pub_url, "", "", "", "", "", "", str(e)])
                continue

            write_row([
                doi,
                title,
                pub_url,
                detail["title_ru"],
                detail["title_en"],
                detail["abstract_ru"],
                detail["abstract_en"],
                detail.get("pages") or "",
                detail["references_ru"],
                detail["references_en"],
            ])

    logger.info("Результат сохранён: %s", output_path)


if __name__ == "__main__":
    main()
